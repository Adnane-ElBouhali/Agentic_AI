# compare_kpi_excel.py

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class Change:
    category: str
    type: str
    detail: str
    old: Any = None
    new: Any = None


def normalize(value: Any) -> Any:
    """Normalize Excel cell values for fair comparison."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def read_sheet(path: Path, sheet_name: str, header_row: int = 1):
    wb = load_workbook(path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")

    ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column

    headers = [
        normalize(ws.cell(header_row, col).value)
        for col in range(1, max_col + 1)
    ]

    rows = []
    for row_idx in range(header_row + 1, max_row + 1):
        row = [
            normalize(ws.cell(row_idx, col).value)
            for col in range(1, max_col + 1)
        ]
        rows.append(row)

    return {
        "path": str(path),
        "sheet": sheet_name,
        "max_row": max_row,
        "max_col": max_col,
        "headers": headers,
        "rows": rows,
    }


def fingerprint_row(row: list[Any]) -> tuple:
    return tuple(row)


def compare_kpi(
    old_file: Path,
    new_file: Path,
    sheet_name: str = "KPI",
    header_row: int = 1,
) -> dict:
    old = read_sheet(old_file, sheet_name, header_row)
    new = read_sheet(new_file, sheet_name, header_row)

    changes: list[Change] = []

    old_headers = old["headers"]
    new_headers = new["headers"]

    old_rows = old["rows"]
    new_rows = new["rows"]

    # 1. Sheet dimensions
    if old["max_row"] != new["max_row"]:
        changes.append(Change(
            category="structure",
            type="row_count_changed",
            detail="Number of rows changed.",
            old=old["max_row"],
            new=new["max_row"],
        ))

    if old["max_col"] != new["max_col"]:
        changes.append(Change(
            category="structure",
            type="column_count_changed",
            detail="Number of columns changed.",
            old=old["max_col"],
            new=new["max_col"],
        ))

    # 2. Column names and order
    old_header_counter = Counter(old_headers)
    new_header_counter = Counter(new_headers)

    deleted_cols = list((old_header_counter - new_header_counter).elements())
    added_cols = list((new_header_counter - old_header_counter).elements())

    for col in deleted_cols:
        changes.append(Change(
            category="columns",
            type="column_deleted",
            detail=f"Column '{col}' was deleted.",
            old=col,
            new=None,
        ))

    for col in added_cols:
        changes.append(Change(
            category="columns",
            type="column_added",
            detail=f"Column '{col}' was added.",
            old=None,
            new=col,
        ))

    if old_header_counter == new_header_counter and old_headers != new_headers:
        changes.append(Change(
            category="columns",
            type="column_order_changed",
            detail="Same column names, but column order changed.",
            old=old_headers,
            new=new_headers,
        ))

        for col in old_header_counter:
            old_positions = [i + 1 for i, h in enumerate(old_headers) if h == col]
            new_positions = [i + 1 for i, h in enumerate(new_headers) if h == col]
            if old_positions != new_positions:
                changes.append(Change(
                    category="columns",
                    type="column_moved",
                    detail=f"Column '{col}' moved.",
                    old=old_positions,
                    new=new_positions,
                ))

    if old_headers == new_headers:
        changes.append(Change(
            category="columns",
            type="same_column_names_and_order",
            detail="Column names and order are identical.",
            old=old_headers,
            new=new_headers,
        ))

    # 3. Row deletion/addition/order using full-row fingerprints
    old_fps = [fingerprint_row(r) for r in old_rows]
    new_fps = [fingerprint_row(r) for r in new_rows]

    old_fp_counter = Counter(old_fps)
    new_fp_counter = Counter(new_fps)

    deleted_rows = list((old_fp_counter - new_fp_counter).elements())
    added_rows = list((new_fp_counter - old_fp_counter).elements())

    for row in deleted_rows:
        old_index = old_fps.index(row) + header_row + 1
        changes.append(Change(
            category="rows",
            type="row_deleted",
            detail=f"Row {old_index} was deleted.",
            old={"excel_row": old_index, "values": list(row)},
            new=None,
        ))

    for row in added_rows:
        new_index = new_fps.index(row) + header_row + 1
        changes.append(Change(
            category="rows",
            type="row_added",
            detail=f"Row {new_index} was added.",
            old=None,
            new={"excel_row": new_index, "values": list(row)},
        ))

    if old_fp_counter == new_fp_counter and old_fps != new_fps:
        changes.append(Change(
            category="rows",
            type="row_order_changed",
            detail="Same rows exist, but row order changed.",
            old=None,
            new=None,
        ))

        old_positions = defaultdict(list)
        new_positions = defaultdict(list)

        for i, fp in enumerate(old_fps, start=header_row + 1):
            old_positions[fp].append(i)

        for i, fp in enumerate(new_fps, start=header_row + 1):
            new_positions[fp].append(i)

        for fp in old_positions:
            if old_positions[fp] != new_positions[fp]:
                changes.append(Change(
                    category="rows",
                    type="row_moved",
                    detail="A row moved position.",
                    old={"excel_rows": old_positions[fp], "values": list(fp)},
                    new={"excel_rows": new_positions[fp], "values": list(fp)},
                ))

    if old_fp_counter == new_fp_counter and old_fps == new_fps:
        changes.append(Change(
            category="rows",
            type="same_rows_and_order",
            detail="Rows and row order are identical.",
        ))

    # 4. Cell-by-cell value changes when dimensions match
    if old["max_row"] == new["max_row"] and old["max_col"] == new["max_col"]:
        for r_idx, (old_row, new_row) in enumerate(
            zip(old_rows, new_rows),
            start=header_row + 1,
        ):
            for c_idx, (old_val, new_val) in enumerate(
                zip(old_row, new_row),
                start=1,
            ):
                if old_val != new_val:
                    column_name = old_headers[c_idx - 1] if c_idx <= len(old_headers) else None

                    changes.append(Change(
                        category="cells",
                        type="cell_value_changed",
                        detail=f"Cell changed at row {r_idx}, column {c_idx} ({column_name}).",
                        old=old_val,
                        new=new_val,
                    ))

    modified = any(
        c.type not in {
            "same_column_names_and_order",
            "same_rows_and_order",
        }
        for c in changes
    )

    return {
        "old_file": str(old_file),
        "new_file": str(new_file),
        "sheet": sheet_name,
        "modified": modified,
        "same_number_of_rows": old["max_row"] == new["max_row"],
        "same_number_of_columns": old["max_col"] == new["max_col"],
        "same_column_names": Counter(old_headers) == Counter(new_headers),
        "same_column_order": old_headers == new_headers,
        "same_rows": Counter(old_fps) == Counter(new_fps),
        "same_row_order": old_fps == new_fps,
        "changes": [asdict(c) for c in changes],
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("old_file", type=Path)
    parser.add_argument("new_file", type=Path)
    parser.add_argument("--sheet", default="KPI")
    parser.add_argument("--header-row", type=int, default=1)
    parser.add_argument("--output", type=Path, default=Path("kpi_changelog.json"))

    args = parser.parse_args()

    result = compare_kpi(
        old_file=args.old_file,
        new_file=args.new_file,
        sheet_name=args.sheet,
        header_row=args.header_row,
    )

    args.output.write_text(
        json.dumps(result, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    print(f"Modified: {result['modified']}")
    print(f"Changelog written to: {args.output}")


if __name__ == "__main__":
    main()